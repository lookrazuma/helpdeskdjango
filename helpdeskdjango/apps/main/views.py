from django.http import HttpResponse, Http404, HttpResponseRedirect, HttpResponseNotFound
from django.shortcuts import render, redirect, reverse
from .models import Task, status, Profile
from .forms import TaskForm, UserUpdateForm, ProfileUpdateForm, TaskContactorForm
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.views.generic import View, ListView
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Count
from django.views.generic import View, ListView, UpdateView
from django.contrib.auth.models import User

from .utils import ObjUpdateMixin, ObjDeleteMixin
from django.db import transaction

from .resources import TaskResource
from tablib import Dataset

from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import serializers
from rest_framework.viewsets import ModelViewSet
from .serializers import TaskSerializer

from import_export import widgets
from import_export.widgets import ManyToManyWidget, ForeignKeyWidget, BooleanWidget
from import_export.forms import ImportForm, ConfirmImportForm
from datetime import datetime, timezone
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from openpyxl.chart import AreaChart, Reference, Series, BarChart # Графики в Excel. Не интегрированы
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill # Стили, применяемые к экспортируемой таблице

from import_export.fields import Field
from dateutil.relativedelta import relativedelta
from django.db.models import Sum, Avg
# from django.shortcuts import render_to_response
from django.template import RequestContext
from qsstats import QuerySetStats
from django.conf import settings
from django.core.mail import send_mail, EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags

class StatusTask:
    def get_status(self):
        return Task.objects.values("status_task").distinct()

class TaskView(StatusTask, ListView):
    paginate_by = 5
    model = Task
    def get_queryset(self):
        username = self.request.user
        task = Task.objects.filter(author=username).order_by('-id')
        return task
    template_name ="main/task_list.html"

class FilterTaskView(ListView, StatusTask):
    def get_queryset(self):
        username = self.request.user
        queryset = Task.objects.filter(author=username, status_task__in=self.request.GET.getlist("status_task"))
        return queryset

class TaskUpdate(LoginRequiredMixin,UpdateView):
    model = Task
    form_class = TaskForm
    def get_queryset(self):
        try:
            username = self.request.user
            task = Task.objects.filter(author=username, status_task="Заявка отправлена")
            return task 
        except:
            raise Http404 ("Статья не найдена!") 
             
    template_name ="main/task_update.html"  
    def get_success_url(self):
        return reverse('task_list/')
 
# Главная страница
def main(request):
    return render(request, 'main/index.html')

# Руководство пользователя
def about(request):
    return render(request, 'main/about.html')

def get_context_data(self, *args, **kwargs):
    context = super().get_context_data(*args, **kwargs)
    context["STATUS_CHOICES"] = STATUS_CHOICES.objects.all()
    return context
 

#@permission_required('main.can_mark_returned')
# Класс обновления пользовательской информации через смешение моделей
class UserUpdateView(LoginRequiredMixin,View, ObjUpdateMixin):
    model_form = UserUpdateForm
    data = User
    raise_exeption = True
    template = 'main/user_account.html'
    

    def get(self, request):
        data_obj = self.data.objects.get(username=self.request.user.username)
        bound_form = self.model_form(instance=data_obj)
        return render(request, self.template, context={'form': bound_form, 'obj': data_obj})
    
    def post(self, request):
        data_obj = self.data.objects.get(username=self.request.user.username)
        bound_form = self.model_form(request.POST, instance=data_obj)

        if bound_form.is_valid():
            new_obj = bound_form.save()
            return redirect('profile_detail_url')
        return render(request, self.template, context={'form': bound_form, 'obj': data_obj})

#Функция отображения личного кабинета
@login_required
@transaction.atomic
def update_profile(request):
    if request.method == 'POST':
        user_form = UserUpdateForm(request.POST, instance=request.user)
        profile_form = ProfileUpdateForm(request.POST, instance=request.user.profile)
        if user_form.is_valid():
            user_form.save()
            profile_form.save()
            # messages.success(request, ('Ваш профиль был успешно обновлен!'))
            return redirect('profile')
    else:
        user_form = UserUpdateForm(instance=request.user)
        profile_form = ProfileUpdateForm(instance=request.user.profile)
    return render(request, 'main/user_account.html', {
        'user_form': user_form,
        'profile_form': profile_form
    })    

# Класс, отвечающий за обновление пользовательского профиля
class ProfileUpdateView( LoginRequiredMixin,View, ObjUpdateMixin):
    model_form = ProfileUpdateForm
    data = Profile
    raise_exeption = True
    template = 'main/user_account.html'
    

    def get(self, request):
        datas = self.data.objects.get(username=self.request.user.id)
        bounds = self.model_form(instance=data_obj)
        return render(request, self.template, context={'form_profile': bounds, 'obj': datas})
    
    def post(self, request):
        datas = self.data.objects.get(username=self.request.user.id)
        bounds = self.model_form(request.POST, instance=datas)

        if bounds.is_valid():
            new_obj = bounds.save()
            return redirect('profile_detail_url')
        return render(request, self.template, context={'form_2': bounds, 'obj': datas})

# Функция отображения формы подачи заявки на странице
def task_show(request):
    if request.user.email:
        tasks = Task.objects.filter(author=request.user) 
        #Дополнительная переменная воизбежание ошибок
        error = ''
        #Запись в БД через метод "Пост"
        if request.method == 'POST':
            form = TaskForm(request.POST)
            #Если форма доступа, то записываем имя юзера-автора и сохраняем в БД
            if form.is_valid():
                instance = form.save(commit=False)
                instance.author = request.user
                instance.save()
                form.save()
                data = {
                    'title': form.cleaned_data['title'],
                    'text': form.cleaned_data['task'],
                    'host_name': form.cleaned_data['host_name'],
                    'dept': form.cleaned_data['dept'],
                    'author': request.user.first_name,
                }
                # mail_client = send_mail(
                #     #subjectфвыыфвфывфыв
                #     form.cleaned_data['title'],
                #     #from 
                #     settings.EMAIL_HOST_USER, 
                #     #to
                #     [request.user.email,],
                #     #except 
                #     fail_silently=False
                #     )

                html_content = render_to_string("email_letters/email_tasker.html", data)
                text_content = strip_tags(html_content)
                mail_client = EmailMultiAlternatives(
                    #subject
                    "Ваша заявка отправлена",
                    #content
                    text_content,
                    #from
                    settings.EMAIL_HOST_USER, 
                    #to(rec list)
                    [request.user.email,],
                    #except 
                    # fail_silently=False
                )
                mail_client.attach_alternative(html_content, "text/html")
                mail_client.send()
                if mail_client and request.user.email != null:
                    messages.success(request, 'Ссылка добавлена')
                    #возвращаем переадресацию пользователя на опр. страниц
                    return redirect('/task_show')
                else:
                    messages.error(request, 'Это было неправильно')
            #Если форма отправки некорректная, то
            else:
                messages.error(request, 'Форма была неверной')


    
        #Получение информации из forms.py класса TaskForm
        form = TaskForm(request.POST or None)
        # Выделение статусов из таблицы Задача, при этом только тех статусов, которые пренадлежат пользователю, работающему с заявкой.
        # Помимо этого, исключаем отображение одинаковых статусов
        status = Task.objects.filter(author=request.user).values("status_task").distinct()

        filter_status = Task.objects.filter(status_task__in=request.GET.getlist("status_task"))
            
        context = {
            'form': form,
            'error': error,
            'tasks': tasks,
            'status': status,
            'filter_status': filter_status
        }
        
        # queryset = Task.objects.filter(status_task = 'Заявка отправлена')

        return render(request, 'main/task_show.html', context)
    else:
        return render(request, "exception/task_no_email.html")  

# Класс фильтрации по статусу
class TaskViewSet(ModelViewSet):
    queryset = Task.objects.all()
    serializer_class = TaskSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['status_task']

def import_data(request):
    if request.user.is_staff:
        if request.method == 'POST':
            file_format = request.POST['file-format']
            task_resource = TaskResource()
            dataset = Dataset()
            new_tasks = request.FILES['importData']

            if file_format == 'CSV':
                imported_data = dataset.load(new_tasks.read().decode('utf-8'),format='csv')
                result = task_resource.import_data(dataset, dry_run=True)                                                                 
            elif file_format == 'JSON':
                imported_data = dataset.load(new_tasks.read().decode('utf-8'),format='json')
                result = task_resource.import_data(dataset, dry_run=True)
            elif file_format == 'xls':
                imported_data = dataset.load(new_tasks.read().decode('utf-8'),format='xls')
                result = task_resource.import_data(dataset, dry_run=True)      

            if not result.has_errors():
                # Import now
                task_resource.import_data(dataset, dry_run=False)

        return render(request, 'main/import.html')
    else:
        return HttpResponseNotFound("Похоже, такой страницы не существует")      
from datetime import datetime
# Функция, отвечающая за экспорт всех задач
def export_tasks_to_xlsx(request):
    # Доступно только пользователю с флагом Администратор
    if request.user.is_superuser:
        """
        Загружает все задачи как XLSX документ на одном листе
        """
        task_queryset = Task.objects.all()
        # status_task = Task.objects.values("status_task")
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=Список всех задач от {date}.xlsx'.format(
            date=datetime.now(timezone.utc),
        )
        workbook = Workbook()
        # Получаю активную книгу Excel
        worksheet = workbook.active
        centered_alignment = Alignment(horizontal='center')

        task = Task.objects.all()
        worksheet.tabColor = '8B0000'
        
        # worksheet['H5'] = 'Количество книг:'
        # worksheet['I5'] = len(books)

        # worksheet['H7'] = 'Книг Романов:'
        # worksheet['I7'] = len(books.filter(name='Роман'))
        
        # worksheet['H10'] = 'Книг от автора Данил:'
        # worksheet['I10'] = len(books.filter(author=request.user))
        columns = [
            ('id'),
            ('Дата создания'),
            ('Название'),
            ('Заявитель'),
            ('Хост компьютера'),
            ('Описание'),
            ('Исполнитель'),
            ('Статус'),
            ('Время выполнения')
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            # Замораживаю строку заголовка таблицы
            worksheet.freeze_panes = worksheet['A2']
            # Задаю ширину колонок
            worksheet.column_dimensions["A"].width = 5
            worksheet.column_dimensions["B"].width = 20
            worksheet.column_dimensions["C"].width = 30
            worksheet.column_dimensions["D"].width = 15
            worksheet.column_dimensions["E"].width = 20
            worksheet.column_dimensions["F"].width = 45
            worksheet.column_dimensions["G"].width = 20
            worksheet.column_dimensions["H"].width = 20
            worksheet.column_dimensions["I"].width = 20

            # Стили для 1-й строки-заголовка
            cell.font = Font(name='Calibri', bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                bottom=Side(border_style='medium', color='FF000000'), 
                top=Side(border_style='medium', color='FF000000'), 
                left=Side(border_style='medium', color='FF000000'), 
                right=Side(border_style='medium', color='FF000000'),
            )
        
        font = Font(name='Calibri', bold=True)
        # Итерация всех задач
        
        for task in task_queryset:
            
            row_num += 1 
               #datetime.strftime((task.in_progress_time), "%b %d %Y %H:%M:%S")
            _in_progress_time = task.in_progress_time

            contractor = str(task.contractor)
            if contractor == "None":
                contractor = "Исполнитель ещё не назначен"
            else:
                contractor
            
            # Определяю данные для вывода в каждую строку
            row = [
                task.pk,
                task.create_date.strftime("%d %m %Y %H:%M:%S"),
                task.title,
                str(task.author),
                task.host_name,
                task.task,
                contractor,
                task.status_task,
                #task.in_progress_time
                _in_progress_time.strftime("%d %m %Y %H:%M:%S")
             
            ]

            # 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cellser = worksheet.cell(row=row_num, column=col_num)
                wrapped_alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = Border(
                bottom=Side(border_style='thin', color='FF000000'),
                left=Side(border_style='thin', color='FF000000'), 
                right=Side(border_style='thin', color='FF000000'),)
                cell.alignment = wrapped_alignment    
        workbook.save(response)

        return response
    else:
        return HttpResponseNotFound("Похоже, такой страницы не существует")   

def time_series(queryset, date_field, interval, func=None):
    qsstats = QuerySetStats(queryset, date_field, func)
    return qsstats.time_series(*interval)

def home(request):
    series = {'count': [], 'total': []}
    tasks = Task.objects.all()
    queryset = tasks
    y = 2021
    for m in range(5, 7):
        start = datetime(y, m, 1)
        end = start + relativedelta(months=1)
        series['count'].append(time_series(queryset, 'create_date', [start, end]))
        series['total'].append(time_series(queryset, 'create_date', [start, end], func=Sum('num_for_chart')))

    start = datetime(y, 1, 1)
    end = start + relativedelta(months=3)
    series['count_6'] = time_series(queryset, 'create_date', [start, end])
    series['total_6'] = time_series(queryset, 'create_date', [start, end], func=Avg('num_for_chart'))

    return render(request, 'main/charts.html', {'series': series})  

class TaskContractorView(StatusTask, ListView):
    model = Task
    def get_queryset(self):
        username = self.request.user
        task = Task.objects.filter(contractor=username)
        return task
    template_name = "main/manager.html"

class Task_ContractorUpdate(LoginRequiredMixin,UpdateView):

    model = Task
    form_class = TaskContactorForm
    def get_queryset(self):
        try:
            username = self.request.user
            task = Task.objects.filter(contractor=username)
            return task 
        except:
            raise Http404 ("Статья не найдена!") 
             
    template_name ="main/manager_details.html" 
    def get_success_url(self):
        return reverse('contractor/')


def new_charts(request):
    return render(request, 'main/new_charts.html')

from django.shortcuts import render
from django.db.models import Sum
from django.http import JsonResponse
def contractor_task_count_chart(request):
    labels = []
    data = []
    # myset = set(labels)
    # username = request.user
    queryset = Task.objects.all().values('contractor__username').annotate(num_for_chart=Sum('num_for_chart')).order_by('contractor')
    for entry in queryset:
        labels.append(entry['contractor__username'])
        data.append(entry['num_for_chart']) 
    
    return JsonResponse(data={
        'labels': labels,
        'data': data,
    })

def status_task_count_chart(request):
    labels = []
    data = []
    # myset = set(labels)
    # username = request.user
    queryset = Task.objects.all().values('status_task').annotate(num_for_chart=Sum('num_for_chart')).order_by('status_task')
    for entry in queryset:
        labels.append(entry['status_task'])
        data.append(entry['num_for_chart']) 
    
    return JsonResponse(data={
        'labels': labels,
        'data': data,
    })

def dept_task_count_chart(request):
    labels = []
    data = []
    # myset = set(labels)
    # username = request.user
    queryset = Task.objects.all().values('dept').annotate(num_for_chart=Sum('num_for_chart')).order_by('dept')
    for entry in queryset:
        labels.append(entry['dept'])
        data.append(entry['num_for_chart']) 
    
    return JsonResponse(data={
        'labels': labels,
        'data': data,
    })
# month
def author_task_count_chart(request):
    labels = []
    data = []
    # myset = set(labels)
    # username = request.user
    queryset = Task.objects.all().values('author__username').annotate(num_for_chart=Sum('num_for_chart')).order_by('author__username')
    for entry in queryset:
        labels.append(entry['author__username'])
        data.append(entry['num_for_chart']) 
    
    return JsonResponse(data={
        'labels': labels,
        'data': data,
    })


def month_task_count_chart(request):
    labels = []
    data = []
    # myset = set(labels)
    # username = request.user
    # dates = Task.objects.filter(create)
    queryset = Task.objects.all().values('create_date').annotate(num_for_chart=Sum('num_for_chart')).order_by(chr('create_date').date())
    for entry in queryset:
        labels.append(entry['create_date'])
        data.append(entry['num_for_chart']) 
    
    return JsonResponse(data={
        'labels': labels,
        'data': data,
    })


def user_statistics(request):
    tasks = Task.objects.filter(author=request.user)
    return render(request, 'main/user_statistic.html', {'tasks': tasks})

def user_statistic_chart(request):
    labels = []
    data = []
    # myset = set(labels)
    # username = request.user
    # dates = Task.objects.filter(create)
    

    
    queryset = Task.objects.filter(author=request.user).values('status_task').annotate(num_for_chart=Sum('num_for_chart')).order_by('status_task')
    for entry in queryset:
        labels.append(entry['status_task'])
        data.append(entry['num_for_chart']) 
    
    return JsonResponse(data={
        'labels': labels,
        'data': data,
    })